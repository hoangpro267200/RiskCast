# =============================================================================
# RISKCAST v5.1.5 — Premium Green DARK Edition
# Author: Bùi Xuân Hoàng — Premium Chart Upgrade by Kai
# =============================================================================

import io
import warnings
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional
from enum import Enum

import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from fpdf import FPDF

warnings.filterwarnings("ignore")

# Optional ARIMA
try:
    from statsmodels.tsa.arima.model import ARIMA
    ARIMA_AVAILABLE = True
except ImportError:
    ARIMA_AVAILABLE = False


# =============================================================================
# CONSTANTS & DATA MODELS
# =============================================================================

class CriterionType(Enum):
    COST = "cost"
    BENEFIT = "benefit"


@dataclass
class AnalysisParams:
    cargo_value: float
    good_type: str
    route: str
    method: str
    month: int
    priority: str
    use_fuzzy: bool
    use_arima: bool
    use_mc: bool
    use_var: bool
    mc_runs: int
    fuzzy_uncertainty: float


@dataclass
class AnalysisResult:
    results: pd.DataFrame
    weights: pd.Series
    data_adjusted: pd.DataFrame
    var: Optional[float]
    cvar: Optional[float]
    historical: np.ndarray
    forecast: np.ndarray
    forecast_months: np.ndarray
    fuzzy_table: Optional[pd.DataFrame]


CRITERIA = [
    "C1: Tỷ lệ phí",
    "C2: Thời gian xử lý",
    "C3: Tỷ lệ tổn thất",
    "C4: Hỗ trợ ICC",
    "C5: Chăm sóc KH",
    "C6: Rủi ro khí hậu",
]

DEFAULT_WEIGHTS = np.array([0.20, 0.15, 0.20, 0.20, 0.10, 0.15])

COST_BENEFIT_MAP: Dict[str, CriterionType] = {
    "C1: Tỷ lệ phí": CriterionType.COST,
    "C2: Thời gian xử lý": CriterionType.COST,
    "C3: Tỷ lệ tổn thất": CriterionType.COST,
    "C4: Hỗ trợ ICC": CriterionType.BENEFIT,
    "C5: Chăm sóc KH": CriterionType.BENEFIT,
    "C6: Rủi ro khí hậu": CriterionType.COST,
}

SENSITIVITY_MAP: Dict[str, float] = {
    "Chubb": 0.95,
    "PVI": 1.10,
    "InternationalIns": 1.20,
    "BaoViet": 1.05,
    "Aon": 0.90,
}


# =============================================================================
# CSS — PREMIUM GREEN DARK THEME
# =============================================================================

def apply_custom_css():
    """Premium Green DARK UI v2 — Xanh đậm neon, bo góc đẹp."""
    st.markdown(
        """
        <style>

        /* BACKGROUND ---------------------------------------------------------*/
        .stApp {
            background: linear-gradient(180deg,#002B1A 0%, #001F13 40%, #00150D 100%) !important;
            font-family: 'Inter', sans-serif;
            color: #E6FFE9 !important;
        }

        .block-container {
            padding: 2rem 2.5rem !important;
            background: rgba(0, 50, 30, 0.30);
            backdrop-filter: blur(5px);
            border-radius: 18px;
            border: 1.5px solid #00D387;
            box-shadow: 0 0 25px rgba(0,255,150,0.15);
        }

        h1 {
            color: #00FFAA !important;
            font-weight: 900 !important;
            text-shadow: 0 0 15px rgba(0,255,160,0.6);
        }
        h2, h3 {
            color: #7CFFCE !important;
            font-weight: 800;
            text-shadow: 0 0 10px rgba(0,255,140,0.4);
        }

        p, label, span, .stMarkdown {
            color: #D7FFEE !important;
            font-weight: 600;
        }

        /* SIDEBAR -----------------------------------------------------------*/
        section[data-testid="stSidebar"] {
            background: #001E14 !important;
            border-right: 3px solid #00FFAA;
        }

        section[data-testid="stSidebar"] h2 {
            color: #00FFAA !important;
            background: #003822;
            padding: 12px;
            border-radius: 10px;
            text-align: center;
            font-weight: 900;
            border: 1.3px solid #00D387;
        }

        section[data-testid="stSidebar"] input,
        section[data-testid="stSidebar"] select {
            background: #002C1A !important;
            color: #D0FFE8 !important;
            border: 1.5px solid #00FFAA !important;
            border-radius: 8px !important;
        }

        .stButton > button {
            background: linear-gradient(135deg,#00A86B,#00FFAA) !important;
            color: #002214 !important;
            border-radius: 12px !important;
            font-size: 1.15rem !important;
            font-weight: 900 !important;
            border: none !important;
        }

        </style>
        """,
        unsafe_allow_html=True,
    )
# =============================================================================
# DATA SERVICE (Mock data for demo) — RISKCAST v5.1.5
# =============================================================================

class DataService:
    """Cung cấp dữ liệu lịch sử rủi ro khí hậu (12 tháng)."""

    @staticmethod
    @st.cache_data(ttl=600)
    def load_historical_data() -> pd.DataFrame:
        # 12 tháng rủi ro (0–1)
        data = {
            "VN - EU": [0.20, 0.22, 0.25, 0.28, 0.27, 0.30, 0.33, 0.35, 0.32, 0.28, 0.25, 0.22],
            "VN - US": [0.18, 0.19, 0.21, 0.24, 0.26, 0.29, 0.31, 0.34, 0.33, 0.29, 0.26, 0.22],
            "VN - Singapore": [0.12, 0.13, 0.15, 0.16, 0.14, 0.15, 0.17, 0.18, 0.17, 0.15, 0.14, 0.12],
            "VN - China": [0.14, 0.15, 0.17, 0.18, 0.20, 0.23, 0.25, 0.27, 0.26, 0.23, 0.20, 0.17],
            "Domestic": [0.08, 0.09, 0.10, 0.11, 0.10, 0.11, 0.12, 0.13, 0.12, 0.10, 0.09, 0.08],
        }
        return pd.DataFrame(data)
# =============================================================================
# FUZZY AHP — Premium Green
# =============================================================================

class FuzzyAHP:

    @staticmethod
    def build_fuzzy_table(weights: pd.Series, uncertainty_pct: float) -> pd.DataFrame:
        factor = uncertainty_pct / 100
        base = weights.values

        # raw (chưa normalize)
        low_raw = np.maximum(base * (1 - factor), 1e-9)
        mid_raw = base.copy()
        high_raw = np.minimum(base * (1 + factor), 0.999)

        # normalize
        def normalize(arr):
            s = arr.sum()
            return arr / s if s > 0 else np.full_like(arr, 1/len(arr))

        low = normalize(low_raw)
        mid = normalize(mid_raw)
        high = normalize(high_raw)

        centroid_raw = (low_raw + mid_raw + high_raw) / 3
        centroid = normalize(centroid_raw)

        rng = high - low

        return pd.DataFrame({
            "low": low,
            "mid": mid,
            "high": high,
            "centroid": centroid,
            "range": rng,
        }, index=weights.index)

    @staticmethod
    def apply(weights: pd.Series, uncertainty_pct: float):
        fuzzy_table = FuzzyAHP.build_fuzzy_table(weights, uncertainty_pct)
        centroid_weight = fuzzy_table["centroid"]
        return centroid_weight, fuzzy_table
# =============================================================================
# MONTE CARLO SIMULATOR
# =============================================================================

class MonteCarloSimulator:

    @staticmethod
    @st.cache_data(ttl=600)
    def simulate(base_risk: float, sensitivity_map: Dict[str, float], n_simulations: int):

        rng = np.random.default_rng(2025)
        companies = list(sensitivity_map.keys())

        mu = np.array([base_risk * sensitivity_map[c] for c in companies])
        sigma = np.maximum(0.02, mu * 0.12)

        sims = rng.normal(mu, sigma, size=(n_simulations, len(companies)))
        sims = np.clip(sims, 0.0, 1.0)

        return companies, sims.mean(axis=0), sims.std(axis=0)
# =============================================================================
# TOPSIS ANALYZER
# =============================================================================

class TOPSISAnalyzer:

    @staticmethod
    def analyze(data: pd.DataFrame, weights: pd.Series, cost_benefit: Dict[str, CriterionType]):
        M = data[list(weights.index)].values.astype(float)

        denom = np.sqrt((M ** 2).sum(axis=0))
        denom[denom == 0] = 1.0
        R = M / denom

        V = R * weights.values

        is_cost = np.array([cost_benefit[c] == CriterionType.COST for c in weights.index])

        ideal_best = np.where(is_cost, V.min(axis=0), V.max(axis=0))
        ideal_worst = np.where(is_cost, V.max(axis=0), V.min(axis=0))

        d_plus = np.sqrt(((V - ideal_best) ** 2).sum(axis=1))
        d_minus = np.sqrt(((V - ideal_worst) ** 2).sum(axis=1))

        return d_minus / (d_plus + d_minus + 1e-12)
# =============================================================================
# RISK CALCULATOR
# =============================================================================

class RiskCalculator:

    @staticmethod
    def calculate_var_cvar(loss_rates: np.ndarray, cargo_value: float, confidence=0.95):

        if len(loss_rates) == 0:
            return 0.0, 0.0

        losses = loss_rates * cargo_value
        var = float(np.percentile(losses, confidence * 100))
        tail = losses[losses >= var]

        cvar = float(tail.mean()) if len(tail) > 0 else var
        return var, cvar

    @staticmethod
    def calculate_confidence(results: pd.DataFrame, data: pd.DataFrame):

        eps = 1e-9
        cv_c6 = results["C6_std"] / (results["C6_mean"] + eps)

        conf_c6 = 1 / (1 + cv_c6)
        conf_c6 = 0.3 + 0.7 * (conf_c6 - conf_c6.min()) / (np.ptp(conf_c6) + eps)

        crit_cv = data.std(axis=1) / (data.mean(axis=1) + eps)
        conf_crit = 1 / (1 + crit_cv)
        conf_crit = 0.3 + 0.7 * (conf_crit - conf_crit.min()) / (np.ptp(conf_crit) + eps)

        return np.sqrt(conf_c6 * conf_crit)
# =============================================================================
# FORECASTER — Only 1-month forecast (fix jumping month)
# =============================================================================

class Forecaster:

    @staticmethod
    def forecast(historical: pd.DataFrame, route: str, current_month: int,
                 months_ahead=1, use_arima=True):

        if route not in historical.columns:
            route = historical.columns[0]

        series = historical[route].values  # length = 12

        # ARIMA or fallback
        if use_arima and ARIMA_AVAILABLE and len(series) >= 6:
            try:
                model = ARIMA(series, order=(1,1,1))
                fitted = model.fit()
                fc = fitted.forecast(months_ahead)
                fc = np.clip(fc, 0, 1)
            except:
                fc = np.array([series[-1]])
        else:
            if len(series) >= 3:
                trend = (series[-1] - series[-3]) / 3
            else:
                trend = 0
            fc = np.array([np.clip(series[-1] + trend, 0, 1)])

        next_month = (current_month % 12) + 1
        return series, fc, np.array([next_month])
# =============================================================================
# ANALYSIS CONTROLLER — RISKCAST v5.1.5
# =============================================================================

class AnalysisController:
    """Điều phối toàn bộ pipeline phân tích RISKCAST."""

    def __init__(self):
        self.fuzzy_ahp = FuzzyAHP()
        self.mc = MonteCarloSimulator()
        self.topsis = TOPSISAnalyzer()
        self.riskcalc = RiskCalculator()
        self.forecaster = Forecaster()

    # -------------------------------------------------------------------------
    # MAIN PIPELINE
    # -------------------------------------------------------------------------
    def run_analysis(self, params: AnalysisParams, historical: pd.DataFrame) -> AnalysisResult:

        # 1. Load company data
        company_data = pd.DataFrame({
            "Company": ["Chubb", "PVI", "InternationalIns", "BaoViet", "Aon"],
            "C1: Tỷ lệ phí": [0.30, 0.28, 0.26, 0.32, 0.24],
            "C2: Thời gian xử lý": [6, 5, 8, 7, 4],
            "C3: Tỷ lệ tổn thất": [0.08, 0.06, 0.09, 0.10, 0.07],
            "C4: Hỗ trợ ICC": [9, 8, 6, 9, 7],
            "C5: Chăm sóc KH": [9, 8, 5, 7, 6],
        }).set_index("Company")

        # ---------------------------------------------------------------------
        # 2. Climate base risk (C6)
        # ---------------------------------------------------------------------
        series_12 = historical[params.route].values
        base_risk = float(series_12[params.month - 1])  # tháng 1 → index 0

        # ---------------------------------------------------------------------
        # 3. Monte Carlo (mean + std) → C6
        # ---------------------------------------------------------------------
        if params.use_mc:
            companies, mc_mean, mc_std = self.mc.simulate(
                base_risk,
                SENSITIVITY_MAP,
                params.mc_runs
            )
            order = [companies.index(c) for c in company_data.index]
            mc_mean = mc_mean[order]
            mc_std = mc_std[order]
        else:
            mc_mean = np.zeros(len(company_data))
            mc_std = np.zeros(len(company_data))

        # ---------------------------------------------------------------------
        # 4. Adjust company data
        # ---------------------------------------------------------------------
        data_adj = company_data.copy()
        data_adj["C6: Rủi ro khí hậu"] = mc_mean

        # Hàng giá trị cao → phí tăng nhẹ
        if params.cargo_value > 50000:
            data_adj["C1: Tỷ lệ phí"] *= 1.1

        # ---------------------------------------------------------------------
        # 5. Lấy trọng số (từ sidebar)
        # ---------------------------------------------------------------------
        base_weights = pd.Series(st.session_state["weights"], index=CRITERIA)

        fuzzy_table = None
        if params.use_fuzzy:
            weights, fuzzy_table = self.fuzzy_ahp.apply(base_weights, params.fuzzy_uncertainty)
        else:
            weights = base_weights.copy()

        # ---------------------------------------------------------------------
        # 6. TOPSIS
        # ---------------------------------------------------------------------
        scores = self.topsis.analyze(data_adj, weights, COST_BENEFIT_MAP)

        results = pd.DataFrame({
            "company": data_adj.index,
            "score": scores,
            "C6_mean": mc_mean,
            "C6_std": mc_std,
        }).sort_values("score", ascending=False).reset_index(drop=True)

        results["rank"] = results.index + 1
        results["recommend_icc"] = results["score"].apply(
            lambda s: "ICC A" if s >= 0.75 else ("ICC B" if s >= 0.5 else "ICC C")
        )

        # ---------------------------------------------------------------------
        # 7. CONFIDENCE SCORE
        # ---------------------------------------------------------------------
        conf = self.riskcalc.calculate_confidence(results, data_adj)
        order_map = {comp: conf[i] for i, comp in enumerate(data_adj.index)}
        results["confidence"] = results["company"].map(order_map).round(3)

        # ---------------------------------------------------------------------
        # 8. VaR / CVaR
        # ---------------------------------------------------------------------
        var = cvar = None
        if params.use_var:
            var, cvar = self.riskcalc.calculate_var_cvar(
                results["C6_mean"].values,
                params.cargo_value,
            )

        # ---------------------------------------------------------------------
        # 9. FORECASTER (chỉ dự báo 1 tháng)
        # ---------------------------------------------------------------------
        hist_series, fc_values, fc_months = self.forecaster.forecast(
            historical,
            params.route,
            current_month=params.month,
            months_ahead=1,
            use_arima=params.use_arima,
        )

        # ---------------------------------------------------------------------
        # 10. RETURN FULL RESULT OBJECT
        # ---------------------------------------------------------------------
        return AnalysisResult(
            results=results,
            weights=weights,
            data_adjusted=data_adj,
            var=var,
            cvar=cvar,
            historical=hist_series,
            forecast=fc_values,
            forecast_months=fc_months,
            fuzzy_table=fuzzy_table,
        )
# =============================================================================
# CHART FACTORY — PREMIUM GREEN DARK
# =============================================================================

class ChartFactory:
    """Toàn bộ chart style Premium Green DARK Neon."""

    # -------------------------------------------------------------------------
    # APPLY PREMIUM THEME
    # -------------------------------------------------------------------------
    @staticmethod
    def _apply_theme(fig: go.Figure, title: str) -> go.Figure:
        fig.update_layout(
            template="plotly_dark",
            title=dict(
                text=f"<b>{title}</b>",
                font=dict(size=24, color="#00FFAA", family="Arial Black"),
                x=0.5,
            ),
            font=dict(size=15, color="#CFFFF0", family="Inter"),
            paper_bgcolor="#002016",
            plot_bgcolor="#001A12",
            margin=dict(l=60, r=40, t=80, b=70),
            legend=dict(
                bgcolor="rgba(0,0,0,0.30)",
                bordercolor="#00FFAA",
                borderwidth=1.6,
                font=dict(size=13, color="#CFFFF0"),
            )
        )

        fig.update_xaxes(
            showgrid=True,
            gridcolor="rgba(0,255,160,0.08)",
            linecolor="#00FFAA",
            linewidth=1.3,
            zeroline=False,
        )

        fig.update_yaxes(
            showgrid=True,
            gridcolor="rgba(0,255,160,0.08)",
            linecolor="#00FFAA",
            linewidth=1.3,
            zeroline=False,
        )
        return fig

    # -------------------------------------------------------------------------
    # PIE CHART — Trọng số
    # -------------------------------------------------------------------------
    @staticmethod
    def create_weights_pie(weights: pd.Series, title: str) -> go.Figure:
        colors = ["#00FFAA", "#00D387", "#00A86B", "#008F5D", "#66FFCA", "#33FFB2"]

        fig = go.Figure(
            data=[
                go.Pie(
                    labels=weights.index,
                    values=weights.values,
                    hole=0.45,
                    marker=dict(colors=colors, line=dict(color="#001A12", width=3)),
                    textinfo="percent",
                    textfont=dict(size=15, color="#001A12"),
                    pull=[0.03] * len(weights),
                )
            ]
        )

        fig.update_layout(
            paper_bgcolor="#002016",
            title=dict(text=f"<b>{title}</b>", font=dict(size=22, color="#00FFAA"), x=0.5)
        )
        return fig

    # -------------------------------------------------------------------------
    # TOPSIS Bar Chart
    # -------------------------------------------------------------------------
    @staticmethod
    def create_topsis_bar(results: pd.DataFrame) -> go.Figure:
        df = results.sort_values("score", ascending=True)

        fig = go.Figure(
            data=[
                go.Bar(
                    x=df["score"],
                    y=df["company"],
                    orientation="h",
                    text=df["score"].apply(lambda x: f"{x:.3f}"),
                    textposition="outside",
                    marker=dict(
                        color=df["score"],
                        colorscale=[
                            [0.0, "#004D36"],
                            [0.25, "#00A06B"],
                            [0.50, "#00D387"],
                            [0.75, "#00FFAA"],
                            [1.00, "#66FFD1"],
                        ],
                        line=dict(color="#00FFAA", width=1.8),
                    )
                )
            ]
        )

        fig = ChartFactory._apply_theme(fig, "🏆 TOPSIS Ranking")
        fig.update_xaxes(range=[0, 1], title="Điểm TOPSIS")
        return fig

    # -------------------------------------------------------------------------
    # FORECAST — 1 tháng
    # -------------------------------------------------------------------------
    @staticmethod
    def create_forecast_chart(historical, forecast, forecast_months, route: str):

        months_hist = list(range(1, len(historical) + 1))
        months_fc = list(forecast_months)

        fig = go.Figure()

        # Lịch sử
        fig.add_trace(
            go.Scatter(
                x=months_hist, y=historical,
                mode="lines+markers",
                name="📈 Lịch sử",
                line=dict(color="#00FFAA", width=3),
                marker=dict(color="#00D387", size=9),
            )
        )

        # Dự báo
        fig.add_trace(
            go.Scatter(
                x=months_fc, y=forecast,
                mode="lines+markers",
                name="🔮 Dự báo",
                line=dict(color="#FFB44C", width=3, dash="dash"),
                marker=dict(color="#FF9800", size=11, symbol="diamond"),
            )
        )

        fig = ChartFactory._apply_theme(fig, f"📊 Forecast — {route}")
        fig.update_xaxes(title="Tháng", dtick=1, tickmode="linear")
        fig.update_yaxes(title="Mức rủi ro", tickformat=".0%")
        return fig

    # -------------------------------------------------------------------------
    # FUZZY HEATMAP
    # -------------------------------------------------------------------------
    @staticmethod
    def create_fuzzy_heatmap(fuzzy_table: pd.DataFrame):
        data = fuzzy_table[["low", "mid", "high", "centroid"]].values
        z_text = np.round(data, 3).astype(str)

        fig = px.imshow(
            data,
            x=["Low", "Mid", "High", "Centroid"],
            y=fuzzy_table.index,
            text=z_text,
            aspect="auto",
            color_continuous_scale=[
                (0.0, "#002015"),
                (0.2, "#005A3A"),
                (0.4, "#00A06B"),
                (0.6, "#00D387"),
                (0.8, "#00FFAA"),
                (1.0, "#66FFD1"),
            ],
        )
        fig.update_traces(texttemplate="%{text}", textfont=dict(size=11, color="#001A12"))

        fig = ChartFactory._apply_theme(fig, "🌿 Fuzzy Heatmap")
        return fig

    # -------------------------------------------------------------------------
    # FUZZY TRIANGLE
    # -------------------------------------------------------------------------
    @staticmethod
    def create_fuzzy_triangle(fuzzy_table: pd.DataFrame, criterion: str):

        row = fuzzy_table.loc[criterion]
        x = [row["low"], row["mid"], row["high"]]
        y = [0, 1, 0]

        fig = go.Figure()
        fig.add_trace(
            go.Scatter(
                x=x, y=y,
                mode="lines+markers",
                fill="tozeroy",
                line=dict(color="#00FFAA", width=3),
                marker=dict(size=10, color="#00D387"),
            )
        )

        fig = ChartFactory._apply_theme(fig, f"🔺 Fuzzy Triangle — {criterion}")
        fig.update_xaxes(title="Trọng số", range=[0, max(0.3, row["high"] + 0.05)])
        fig.update_yaxes(title="Membership", range=[0, 1.2])
        return fig


# =============================================================================
# REPORT GENERATOR — EXPORT EXCEL + PDF
# =============================================================================

class ReportGenerator:

    # -------------------------------------------------------------------------
    # EXCEL EXPORT
    # -------------------------------------------------------------------------
    def generate_excel(self, results, data_adj, weights, fuzzy_table):
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Results"

        for r in dataframe_to_rows(results, index=False, header=True):
            ws1.append(r)

        ws2 = wb.create_sheet("Adjusted Data")
        for r in dataframe_to_rows(data_adj, index=True, header=True):
            ws2.append(r)

        ws3 = wb.create_sheet("Weights")
        df_w = pd.DataFrame({"Criterion": weights.index, "Weight": weights.values})
        for r in dataframe_to_rows(df_w, index=False, header=True):
            ws3.append(r)

        if fuzzy_table is not None:
            ws4 = wb.create_sheet("Fuzzy Table")
            for r in dataframe_to_rows(fuzzy_table, index=True, header=True):
                ws4.append(r)

        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    # -------------------------------------------------------------------------
    # PDF EXPORT
    # -------------------------------------------------------------------------
    def generate_pdf(self, results, params, var, cvar):
        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", "B", 16)
            pdf.cell(0, 10, "RISKCAST Summary Report", ln=True)

            pdf.set_font("Arial", "", 12)
            top = results.iloc[0]

            pdf.ln(5)
            pdf.cell(0, 8, f"Recommended Company: {top['company']}", ln=True)
            pdf.cell(0, 8, f"TOPSIS Score: {top['score']:.3f}", ln=True)
            pdf.cell(0, 8, f"Confidence: {top['confidence']:.2f}", ln=True)
            pdf.cell(0, 8, f"ICC Recommendation: {top['recommend_icc']}", ln=True)

            pdf.ln(5)
            pdf.cell(0, 8, f"Route: {params.route}", ln=True)
            pdf.cell(0, 8, f"Cargo Value: ${params.cargo_value:,.0f}", ln=True)

            if var is not None:
                pdf.ln(5)
                pdf.cell(0, 8, f"VaR 95%: ${var:,.0f}", ln=True)
                pdf.cell(0, 8, f"CVaR 95%: ${cvar:,.0f}", ln=True)

            return pdf.output(dest="S").encode("latin-1")

        except Exception:
            return None
# =============================================================================
# STREAMLIT UI — RISKCAST v5.1.5 PREMIUM GREEN DARK
# =============================================================================

class StreamlitUI:
    """Quản lý toàn bộ UI frontend của RISKCAST."""

    def __init__(self):
        self.controller = AnalysisController()
        self.chart = ChartFactory()
        self.exporter = ReportGenerator()

    # -------------------------------------------------------------------------
    # INIT
    # -------------------------------------------------------------------------
    def initialize(self):
        st.set_page_config(
            page_title="RISKCAST v5.1.5 — ESG Logistics Risk Assessment",
            page_icon="🛡️",
            layout="wide",
        )
        apply_custom_css()

        if "weights" not in st.session_state:
            st.session_state["weights"] = DEFAULT_WEIGHTS.copy()

        if "locked" not in st.session_state:
            st.session_state["locked"] = [False] * len(CRITERIA)

    # -------------------------------------------------------------------------
    # SIDEBAR INPUTS
    # -------------------------------------------------------------------------
    def render_sidebar(self) -> AnalysisParams:
        with st.sidebar:
            st.header("📦 Thông tin lô hàng")

            cargo_value = st.number_input(
                "Giá trị lô hàng (USD)", min_value=1000, value=38000, step=1000
            )
            good_type = st.selectbox(
                "Loại hàng",
                ["Điện tử", "Đông lạnh", "Hàng khô", "Hàng nguy hiểm", "Khác"],
            )
            route = st.selectbox(
                "Tuyến vận tải",
                ["VN - EU", "VN - US", "VN - Singapore", "VN - China", "Domestic"],
            )
            method = st.selectbox("Phương thức", ["Sea", "Air", "Truck"])
            month = st.selectbox("Tháng hiện tại", list(range(1, 12 + 1)), index=8)
            priority = st.selectbox(
                "Ưu tiên của chủ hàng",
                ["An toàn tối đa", "Cân bằng", "Tối ưu chi phí"],
            )

            st.markdown("---")
            st.header("⚙️ Cấu hình mô hình")

            use_fuzzy = st.checkbox("Fuzzy AHP", True)
            use_arima = st.checkbox("ARIMA dự báo khí hậu", True)
            use_mc = st.checkbox("Monte Carlo", True)
            use_var = st.checkbox("Tính VaR / CVaR", True)

            mc_runs = st.number_input(
                "Số vòng mô phỏng Monte Carlo",
                min_value=500, max_value=20000, value=2000, step=500,
            )

            fuzzy_uncertainty = st.slider(
                "Mức bất định Fuzzy (%)",
                0, 50, 15
            ) if use_fuzzy else 15

            return AnalysisParams(
                cargo_value=cargo_value,
                good_type=good_type,
                route=route,
                method=method,
                month=month,
                priority=priority,
                use_fuzzy=use_fuzzy,
                use_arima=use_arima,
                use_mc=use_mc,
                use_var=use_var,
                mc_runs=mc_runs,
                fuzzy_uncertainty=fuzzy_uncertainty,
            )

    # -------------------------------------------------------------------------
    # WEIGHT SLIDERS
    # -------------------------------------------------------------------------
    def render_weight_controls(self):
        st.subheader("⚖️ Trọng số tiêu chí (Auto-balance)")

        cols = st.columns(len(CRITERIA))
        new_weights = st.session_state["weights"].copy()

        for i, crit in enumerate(CRITERIA):
            with cols[i]:
                st.markdown(f"**{crit.split(':')[0]}**")

                locked = st.checkbox(
                    "🔒", value=st.session_state["locked"][i], key=f"lock_{i}"
                )
                st.session_state["locked"][i] = locked

                new_w = st.number_input(
                    "Tỉ lệ",
                    min_value=0.0,
                    max_value=1.0,
                    step=0.01,
                    value=float(new_weights[i]),
                    key=f"w_{i}",
                    label_visibility="collapsed",
                )
                new_weights[i] = new_w

                st.write(f"**{new_w:.1%}**")

        # Auto normalize
        st.session_state["weights"] = WeightManager.auto_balance(
            new_weights, st.session_state["locked"]
        )

        total = st.session_state["weights"].sum()
        if abs(total - 1.0) > 0.01:
            st.warning(f"Tổng hiện tại: {total:.1%}")
        else:
            st.success(f"✨ Tổng trọng số = 100%")

    # -------------------------------------------------------------------------
    # DISPLAY RESULTS
    # -------------------------------------------------------------------------
    def display_results(self, result: AnalysisResult, params: AnalysisParams):
        st.success("🎉 Hoàn tất phân tích!")

        # 1. TABLE
        st.subheader("🏅 Kết quả TOPSIS")
        df_view = result.results[
            ["rank", "company", "score", "confidence", "recommend_icc"]
        ].set_index("rank")

        df_view.columns = [
            "Công ty",
            "Điểm TOPSIS",
            "Độ tin cậy",
            "ICC đề xuất",
        ]
        st.dataframe(df_view, use_container_width=True)

        # BEST
        best = result.results.iloc[0]
        st.markdown(
            f"""
            <div class="result-box">
                🏆 <b>Gợi ý tốt nhất: {best['company']}</b><br><br>
                Score: {best['score']:.3f} |
                Confidence: {best['confidence']:.2f} |
                {best['recommend_icc']}
            </div>
            """,
            unsafe_allow_html=True,
        )

        # 2. WEIGHTS PIE
        st.subheader("📊 Trọng số cuối cùng")
        pie = self.chart.create_weights_pie(result.weights, "Final Weights")
        st.plotly_chart(pie, use_container_width=True)

        # 3. TOPSIS CHART
        st.subheader("📈 Biểu đồ TOPSIS")
        fig_top = self.chart.create_topsis_bar(result.results)
        st.plotly_chart(fig_top, use_container_width=True)

        # 4. CLIMATE FORECAST
        st.subheader("🌦️ Dự báo rủi ro khí hậu")
        fig_fc = self.chart.create_forecast_chart(
            result.historical, result.forecast, result.forecast_months, params.route
        )
        st.plotly_chart(fig_fc, use_container_width=True)

        # 5. FUZZY
        if result.fuzzy_table is not None:
            st.subheader("🌿 Fuzzy AHP")

            display = result.fuzzy_table.copy()
            display.columns = ["Low", "Mid", "High", "Centroid", "Range"]

            st.dataframe(display.style.format("{:.3f}"), use_container_width=True)

            # Highlight biggest uncertainty
            strongest = display["Range"].idxmax()
            st.info(f"🔥 Tiêu chí bất định mạnh nhất: **{strongest}**")

            heat = self.chart.create_fuzzy_heatmap(result.fuzzy_table)
            st.plotly_chart(heat, use_container_width=True)

            crit_sel = st.selectbox("Xem Fuzzy Triangle cho tiêu chí:", display.index)
            tri = self.chart.create_fuzzy_triangle(result.fuzzy_table, crit_sel)
            st.plotly_chart(tri, use_container_width=True)

        # 6. ESG FINANCE RISK
        if params.use_var:
            st.subheader("💰 VaR & CVaR")
            st.metric("VaR 95%", f"${result.var:,.0f}")
            st.metric("CVaR 95%", f"${result.cvar:,.0f}")

        # 7. EXPORT
        st.subheader("📥 Xuất báo cáo")

        col1, col2 = st.columns(2)

        with col1:
            excel = self.exporter.generate_excel(
                result.results,
                result.data_adjusted,
                result.weights,
                result.fuzzy_table,
            )
            st.download_button(
                "📊 Tải Excel đầy đủ",
                data=excel,
                file_name="riskcast_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with col2:
            pdf = self.exporter.generate_pdf(
                result.results,
                params,
                result.var,
                result.cvar,
            )
            if pdf:
                st.download_button(
                    "📄 Tải PDF Executive Summary",
                    data=pdf,
                    file_name="riskcast_summary.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )

    # -------------------------------------------------------------------------
    # RUN
    # -------------------------------------------------------------------------
    def run(self):
        self.initialize()

        st.title("🚢 RISKCAST v5.1.5 — ESG Logistics Risk Assessment")
        st.write("Fuzzy AHP • Monte Carlo • VaR & CVaR • Climate Forecast • TOPSIS")

        st.markdown("---")

        historical = DataService.load_historical_data()
        params = self.render_sidebar()

        st.markdown("---")
        self.render_weight_controls()

        st.markdown("---")
        if st.button("🚀 PHÂN TÍCH NGAY", use_container_width=True):
            with st.spinner("Đang xử lý..."):
                result = self.controller.run_analysis(params, historical)
                self.display_results(result, params)


# =============================================================================
# MAIN
# =============================================================================

def main():
    app = StreamlitUI()
    app.run()

if __name__ == "__main__":
    main()
